# -*- coding: utf-8 -*-
"""
test_io_xlsx.py ― io_xlsx.py の round-trip テスト（unittest + openpyxl）

実行: python3 02_第2弾_G条件/webapp/test_io_xlsx.py
検証:
  - 合成データ: セル内改行・末尾空セル(位置)・数値様文字列("007")の保持
  - 実データ: 06土のうの G条件CSV / TC CSV を CSV→xlsx→CSV して正規化一致
  - CSV書き戻しがエンジン形式(cp932/CRLF/QUOTE_MINIMAL)であること
正規化: 全行を最大列数まで '' でパディングし、セル内改行を \n に統一して比較
        （ragged↔矩形の差だけを吸収し、値・位置は厳密に比較する）
"""

import os
import glob
import tempfile
import unittest

import io_xlsx


def _pad(matrix):
    w = max((len(r) for r in matrix), default=0)
    out = []
    for r in matrix:
        row = [(_nl(c)) for c in r] + [''] * (w - len(r))
        out.append(row)
    # 末尾全空行は無視（xlsx使用範囲の揺れ吸収）
    while out and all(c == '' for c in out[-1]):
        out.pop()
    return out


def _nl(s):
    return s.replace('\r\n', '\n').replace('\r', '\n')


class TestSynthetic(unittest.TestCase):
    def setUp(self):
        self.d = tempfile.mkdtemp(prefix='ioxlsx_')

    def _roundtrip(self, matrix):
        c0 = os.path.join(self.d, 'a.csv')
        x = os.path.join(self.d, 'a.xlsx')
        c1 = os.path.join(self.d, 'b.csv')
        io_xlsx.write_csv_matrix(matrix, c0)
        io_xlsx.csv_to_xlsx(c0, x)
        io_xlsx.xlsx_to_csv(x, c1)
        m1, _ = io_xlsx.read_csv_matrix(c1)
        return m1

    def test_embedded_newline_preserved(self):
        m = [['ID', '説明'], ['1', '一行目\n二行目\n三行目'], ['2', 'ok']]
        m1 = self._roundtrip(m)
        self.assertEqual(_pad(m), _pad(m1))
        # セル内改行が本当に残っているか
        self.assertIn('\n', [r for r in m1 if r and r[0] == '1'][0][1])

    def test_trailing_empty_cells_position(self):
        # G条件型: 末尾空セルは列位置として保持されるべき
        m = [['施工区分', 'G1', 'G2', 'G3'], ['規格名計上', '○', '○', ''], ['', 'a', '', '']]
        m1 = self._roundtrip(m)
        self.assertEqual(_pad(m), _pad(m1))

    def test_numeric_like_string_kept(self):
        m = [['code', 'val'], ['007', '0.28'], ['17', '100']]
        m1 = self._roundtrip(m)
        self.assertEqual(_pad(m), _pad(m1))
        # 桁落ちしていないこと
        self.assertTrue(any(r[0] == '007' for r in m1))

    def test_write_format_is_engine_compatible(self):
        m = [['a', 'b'], ['x', 'y\nz']]
        c1 = os.path.join(self.d, 'fmt.csv')
        io_xlsx.write_csv_matrix(m, c1)
        with open(c1, 'rb') as f:
            raw = f.read()
        self.assertIn(b'\r\n', raw)          # CRLF 行終端
        self.assertIn('"y\nz"'.encode('cp932'), raw)  # セル内改行はquoteされcp932で書かれる


class TestConditionCols(unittest.TestCase):
    def test_tc_condition_cols(self):
        header = ['テストID', 'テスト区分', '作業区分', '作業内容', '材料計上区分',
                  '購入土計上区分(固定)', '土のう種類', '代価表行と数量(数率)',
                  '選択肢の適切さ確認', '規格名計上']
        cols = io_xlsx.detect_tc_condition_cols([header])
        self.assertEqual(cols, [3, 4, 5, 6, 7])  # C〜G

    def test_gjoken_not_colored(self):
        header = ['施工区分/入力条件', 'G1', 'G2']
        self.assertEqual(io_xlsx.detect_tc_condition_cols([header]), [])

    def test_header_fill_applied(self):
        import openpyxl as pyxl
        d = tempfile.mkdtemp(prefix='ioxlsx_fill_')
        m = [['テストID', 'テスト区分', '作業区分', '作業内容', '選択肢の適切さ確認'],
             ['TC-001', '通常', 'a', 'b', 'x']]
        x = os.path.join(d, 'c.xlsx')
        io_xlsx.matrix_to_xlsx(m, x)
        ws = pyxl.load_workbook(x).active
        # C,D(=作業区分,作業内容)が塗られ、A,B,E は塗られない
        self.assertEqual(ws.cell(1, 3).fill.fgColor.rgb[-6:], io_xlsx.CONDITION_HEADER_FILL)
        self.assertEqual(ws.cell(1, 4).fill.fgColor.rgb[-6:], io_xlsx.CONDITION_HEADER_FILL)
        self.assertNotEqual(ws.cell(1, 1).fill.fgColor.rgb[-6:], io_xlsx.CONDITION_HEADER_FILL)
        self.assertNotEqual(ws.cell(1, 5).fill.fgColor.rgb[-6:], io_xlsx.CONDITION_HEADER_FILL)

    def test_changed_cell_fill(self):
        """2件目以降のTC行で、直前行と変わった条件セルだけ CHANGED_CELL_FILL で塗る。"""
        import openpyxl as pyxl
        d = tempfile.mkdtemp(prefix='ioxlsx_chg_')
        m = [['テストID', 'テスト区分', '作業区分', '作業内容', '選択肢の適切さ確認'],
             ['TC-001', '回帰', 'a', 'x', ''],   # 1件目=基準(塗らない)
             ['TC-002', '差分', 'b', 'x', ''],   # 作業区分だけ変化
             ['TC-003', '差分', 'b', '-', '']]   # 作業内容が '-' へ(=選択肢の切替でない→塗らない)
        x = os.path.join(d, 'chg.xlsx')
        io_xlsx.matrix_to_xlsx(m, x)
        ws = pyxl.load_workbook(x).active

        def is_changed(r, c):
            f = ws.cell(r, c).fill
            return bool(f.patternType) and f.fgColor.rgb[-6:] == io_xlsx.CHANGED_CELL_FILL

        # 1件目(row2)は一切塗らない
        self.assertFalse(is_changed(2, 3))
        self.assertFalse(is_changed(2, 4))
        # 2件目(row3): 作業区分(col3)が変化 → 塗る / 作業内容(col4)は不変 → 塗らない
        self.assertTrue(is_changed(3, 3))
        self.assertFalse(is_changed(3, 4))
        # 3件目(row4): 作業区分は不変・作業内容は '-' へ → どちらも塗らない
        self.assertFalse(is_changed(4, 3))
        self.assertFalse(is_changed(4, 4))


class TestRealData(unittest.TestCase):
    """実在する 06土のう の G条件 / TC を round-trip。"""

    BASE = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '運用案件', '06_土のう積工'))

    def _rt(self, csv_path):
        d = tempfile.mkdtemp(prefix='ioxlsx_real_')
        x = os.path.join(d, 'o.xlsx')
        c1 = os.path.join(d, 'o.csv')
        m0, _ = io_xlsx.read_csv_matrix(csv_path)
        io_xlsx.csv_to_xlsx(csv_path, x)
        io_xlsx.xlsx_to_csv(x, c1)
        m1, _ = io_xlsx.read_csv_matrix(c1)
        self.assertEqual(_pad(m0), _pad(m1), 'round-trip mismatch: %s' % csv_path)

    def test_gjoken_roundtrip(self):
        p = os.path.join(self.BASE, '20_叩き台G条件', 'Gaia入力基準表_土のう積工(m2)_叩き台.csv')
        if not os.path.exists(p):
            self.skipTest('実データ未配置: %s' % p)
        self._rt(p)

    def test_tc_roundtrip(self):
        p = os.path.join(self.BASE, '60_改定後TC叩き台', 'step3.0_テストケース_土のう積工.csv')
        if not os.path.exists(p):
            self.skipTest('実データ未配置: %s' % p)
        self._rt(p)


if __name__ == '__main__':
    import sys
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    unittest.main(verbosity=2)
