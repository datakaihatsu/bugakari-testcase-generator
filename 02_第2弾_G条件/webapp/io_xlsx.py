# -*- coding: utf-8 -*-
"""
io_xlsx.py ― CSV⇔xlsx 境界層（人に渡す面だけ xlsx 化）

目的は見栄えでなく【事故防止】:
  セル内改行入りCSVをExcelで開いて保存すると cp932/引用符で壊れる。
  G条件は人が編集して③へ戻す往復があり致命的（設計書 §8）。
  → 人に渡す面(G条件・TC)は xlsx。内部/回帰は CSV のまま。変換はこの1枚に閉じる。

方針:
  - 依存は openpyxl のみ（pure-Python・vendoring可・Q1承認）。Webサーバ/ロジックは標準のまま。
  - CSVの読み書きはエンジン(bugakari_json.write_csv)と同一: cp932 / lineterminator='\r\n' / QUOTE_MINIMAL。
  - xlsxセルは【全て文字列として】格納し、数値化・桁落ち("007"→7 等)を防ぐ。
  - セル内改行は本物の改行として保存し wrap_text で表示（往復事故防止の核）。
  - 位置で列が決まるG条件のため、末尾空セルも列位置として保持する（削らない）。

設計書: 運用化設計書_2026-07-07.md §8。
"""

import csv
import os
import unicodedata

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

CSV_ENCODINGS = ('cp932', 'utf-8-sig', 'utf-8')
WRITE_ENCODING = 'cp932'          # エンジンと同一（Excelで開くため必須）
LINE_TERM = '\r\n'                # bugakari_json.write_csv と同一
MAX_COL_WIDTH = 60                # 単一行セルの列幅上限（表示幅目安）
WRAP_COL_WIDTH = 32               # 折り返しセル(セル内改行あり)の列幅上限＝横に広げず縦に折る

# TC「条件（軸）列」ヘッダーの塗り色（積算シミュレートで選ぶ列。標準単価表に色合わせ）。
# 淡いブルー（Excel定番の帳票色）。標準単価表の実色が分かれば差し替える。
CONDITION_HEADER_FILL = 'DCE6F1'
# TCの2件目以降のパターンで「直前の行と選択肢が変わった条件セル」を塗る色（淡い黄）。
# どの条件を切り替えたパターンかを一目で分かるようにする（2026-07-09 運用者フィードバック）。
CHANGED_CELL_FILL = 'FFF2CC'
# TC末尾の観点/期待列（この手前までが条件列）。先頭2列=テストID/テスト区分はメタ。
_TC_TRAILING_EXACT = ('選択肢の適切さ確認', '規格名計上')
_TC_TRAILING_PREFIX = ('代価表行と数量', '期待:')


# ----------------------------------------------------------------------------
# CSV ⇔ matrix（エンジンと同一のI/O）
# ----------------------------------------------------------------------------
def read_csv_matrix(path):
    """CSV → list[list[str]]。cp932優先で復号。セル内改行は保持（newline=''）。"""
    last_err = None
    for enc in CSV_ENCODINGS:
        try:
            with open(path, 'r', encoding=enc, newline='') as f:
                return [list(row) for row in csv.reader(f)], enc
        except UnicodeDecodeError as e:
            last_err = e
    raise last_err


def write_csv_matrix(matrix, path, encoding=WRITE_ENCODING):
    """matrix → CSV。エンジンと同一設定（cp932 / CRLF / QUOTE_MINIMAL）。"""
    import io
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator=LINE_TERM)
    for row in matrix:
        writer.writerow(row)
    with open(path, 'w', encoding=encoding, newline='') as f:
        f.write(buf.getvalue())


# ----------------------------------------------------------------------------
# matrix ⇔ xlsx
# ----------------------------------------------------------------------------
def _cell_to_str(v):
    """xlsxセル値 → 文字列。人編集で数値化された値も安全に文字列化。"""
    if v is None:
        return ''
    if isinstance(v, bool):
        return '1' if v else '0'
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)
    return str(v)


def _disp_width(s):
    """表示幅の目安（全角/東アジア幅Wide,Fullは2、それ以外1）。列幅算出用。"""
    return sum(2 if unicodedata.east_asian_width(ch) in ('W', 'F') else 1 for ch in s)


def _note_boundary(matrix):
    """G条件の「(注)」行の位置を返す（無ければ末尾）。列幅算出をこの上の表部分に限定する。"""
    for i, row in enumerate(matrix):
        if row and str(row[0]).strip() == '(注)':
            return i
    return len(matrix)


def detect_tc_condition_cols(matrix):
    """TCの「条件（軸）列」の1始まり列番号リストを返す。
    TC判定: ヘッダー先頭が「テストID」。条件列 = 先頭メタ2列(テストID/テスト区分)の後ろ〜
    末尾の観点/期待列(代価表行と数量/選択肢の適切さ確認/規格名計上/期待:)の手前。
    TCでない(例 G条件)なら空リスト。"""
    if not matrix:
        return []
    header = [str(h).strip() for h in matrix[0]]
    if not header or header[0] != 'テストID':
        return []
    end = len(header)
    for i, h in enumerate(header):
        if h in _TC_TRAILING_EXACT or any(h.startswith(p) for p in _TC_TRAILING_PREFIX):
            end = i
            break
    start = 2  # テストID, テスト区分 の後ろ
    if end <= start:
        return []
    return [i + 1 for i in range(start, end)]  # 1始まり列番号


def matrix_to_xlsx(matrix, xlsx_path, sheet_title='Sheet1',
                   highlight_header_cols=None, header_fill=CONDITION_HEADER_FILL):
    """list[list[str]] → xlsx。全セル文字列格納・セル内改行はwrap保存・値のみ（装飾最小）。
    列幅は「(注)」より上の表部分だけで算出する（注の長文で列が広がるのを防ぐ。
    注は右の空セルへオーバーフロー表示される）。
    highlight_header_cols=None のときはTC条件列を自動判定して見出しを塗る
    （TCでなければ塗らない）。空リストを渡せば塗らない。"""
    if highlight_header_cols is None:
        highlight_header_cols = detect_tc_condition_cols(matrix)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (sheet_title or 'Sheet1')[:31]  # Excelのシート名31文字制限
    wrap = Alignment(wrap_text=True, vertical='top')
    topalign = Alignment(vertical='top')  # 非折返しセルもExcel既定(下揃え)でなく上揃えに統一

    note_row = _note_boundary(matrix)  # この行index(0始まり)以降は幅算出・wrap対象から除外
    col_maxline = {}       # 列 -> 最大単一行表示幅（(注)より上）
    col_is_long = {}       # 列 -> 長文列か（改行あり or WRAP_COL_WIDTH超を含む）
    cells_above = {}       # 列 -> (注)より上のセル一覧（長文列に一括wrap適用するため）
    for r, row in enumerate(matrix, 1):
        for c, val in enumerate(row, 1):
            s = _cell_to_str(val)
            if s == '':
                continue  # 空セルは書かない（位置は max_column で保たれる）
            cell = ws.cell(row=r, column=c, value=s)
            cell.number_format = '@'  # 数値/日付への誤変換防止
            has_nl = ('\n' in s or '\r' in s)
            cell.alignment = wrap if has_nl else topalign  # 全セル上揃え（既定の下揃えを回避）
            if (r - 1) < note_row:
                longest = max((_disp_width(x) for x in s.splitlines()), default=_disp_width(s))
                col_maxline[c] = max(col_maxline.get(c, 0), longest)
                if has_nl or longest > WRAP_COL_WIDTH:
                    col_is_long[c] = True
                cells_above.setdefault(c, []).append(cell)

    # 長文列は幅を WRAP_COL_WIDTH で頭打ちにし列全体を折り返す（横に広げず縦に折る）。
    # それ以外は内容幅にフィット（単一行上限 MAX_COL_WIDTH）。
    for c, maxw in col_maxline.items():
        if col_is_long.get(c):
            width = WRAP_COL_WIDTH
            for cell in cells_above[c]:
                cell.alignment = wrap
        else:
            width = min(maxw + 2, MAX_COL_WIDTH)
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = width
    # 先頭行を見出しとして固定（装飾は最小限）
    if matrix:
        for c in range(1, (max((len(r) for r in matrix), default=0)) + 1):
            hc = ws.cell(row=1, column=c)
            if hc.value is not None:
                hc.font = Font(bold=True)
        ws.freeze_panes = 'A2'
        # 条件（軸）列の見出しを塗る（積算シミュレートで選ぶ列＝標準単価表に色合わせ）
        if highlight_header_cols:
            fill = PatternFill('solid', fgColor=header_fill)
            for c in highlight_header_cols:
                ws.cell(row=1, column=c).fill = fill

    # TCの2件目以降のパターン: 直前行と変わった条件セルだけ塗る（変更点の可視化）。
    #   TC（条件列が取れる）ときのみ。'-'/空への変化は「選択肢の切替」ではないので塗らない。
    diff_cols = detect_tc_condition_cols(matrix)
    if diff_cols:
        change_fill = PatternFill('solid', fgColor=CHANGED_CELL_FILL)
        prev = None
        for r in range(2, len(matrix) + 1):      # 1始まり行番号（1=見出し）
            row = matrix[r - 1]
            if prev is not None:                 # 2件目以降のパターンのみ比較
                for c in diff_cols:
                    cur = str(row[c - 1]).strip() if c - 1 < len(row) else ''
                    pv = str(prev[c - 1]).strip() if c - 1 < len(prev) else ''
                    if cur != pv and cur not in ('', '-'):
                        cell = ws.cell(row=r, column=c)
                        if cell.value is not None:
                            cell.fill = change_fill
            prev = row

    os.makedirs(os.path.dirname(os.path.abspath(xlsx_path)), exist_ok=True)
    wb.save(xlsx_path)
    return xlsx_path


def xlsx_to_matrix(xlsx_path, sheet=None):
    """xlsx → list[list[str]]。空セルは''、数値は文字列化。使用範囲(max_row×max_column)で矩形化。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    matrix = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column, values_only=True):
        matrix.append([_cell_to_str(v) for v in row])
    return matrix


# ----------------------------------------------------------------------------
# 高水準API（ファイル→ファイル）
# ----------------------------------------------------------------------------
def csv_to_xlsx(csv_path, xlsx_path, sheet_title=None, header_fill=CONDITION_HEADER_FILL):
    """CSVファイル → xlsxファイル。人に渡す面の生成。
    TCなら条件列見出しを自動で header_fill 色に塗る（G条件は塗らない）。"""
    matrix, _enc = read_csv_matrix(csv_path)
    title = sheet_title or os.path.splitext(os.path.basename(csv_path))[0]
    return matrix_to_xlsx(matrix, xlsx_path, title, header_fill=header_fill)


def xlsx_to_csv(xlsx_path, csv_path, encoding=WRITE_ENCODING):
    """xlsxファイル → CSVファイル。③へ戻す（エンジンが読む形＝cp932/CRLF）。"""
    matrix = xlsx_to_matrix(xlsx_path)
    matrix = _rstrip_trailing_empty_rows(matrix)
    write_csv_matrix(matrix, csv_path, encoding=encoding)
    return csv_path


def _rstrip_trailing_empty_rows(matrix):
    """末尾の全空行を除去（xlsx使用範囲が余分な空行を含む場合の保険）。"""
    m = list(matrix)
    while m and all(c == '' for c in m[-1]):
        m.pop()
    return m


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 3:
        print('Usage: python3 io_xlsx.py <in.csv|in.xlsx> <out.xlsx|out.csv>')
        sys.exit(1)
    src, dst = sys.argv[1], sys.argv[2]
    if src.lower().endswith('.csv'):
        csv_to_xlsx(src, dst)
    else:
        xlsx_to_csv(src, dst)
    print('OK:', dst)
